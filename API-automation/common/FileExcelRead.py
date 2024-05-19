import openpyxl
from common.Data_Encrypt_AES import ENCRYPTAES


from config import *
class FileExcelRead():
    @staticmethod
    #  TODO 读取excel文件,需要知道文件路径，sheet页名
    def read_excel(excel_file_path=EXCEL_FILE_PATH6,sheet_name=SHEET_NAME):
        """
        :param excel_file_path:   excel文件的路径，EXCEL_FILE_PATH是维护在config.py的常量
        :param sheet_name:  excel要操作的sheet页名，SHEET_NAME是维护在config.py的常量
        :return:  返回读取的结果
        """
        try:
            # 判断是否有这个文件，如果文件存在，就加载文件，返回一个workbook对象
            # openpyxl.load_workbook(),用于加载excel文件并且返回一个Workbook对象，在python中对excel文件进行操作
            workbook=openpyxl.load_workbook(excel_file_path)
            print("返回的workbook:",workbook)
        except:
            # 如果文件不存在，就创建一个新的excel对象
            workbook=openpyxl.Workbook()
            raise FileNotFoundError("找不到excel文件")


        # 判断传入的sheet_name是不是在excel文件中,如果存在，就加载该sheet页的数据，work.sheetnames是获取excel的所有sheet页
        if sheet_name in workbook.sheetnames:
            worksheet=workbook[sheet_name]
        else:
            # 如果文件中没有该sheet就创建一个新的sheet页
            worksheet=workbook.create_sheet(sheet_name)

        # todo 获取列名，把sheet页第2行的数据拿到当做key值
        headers=[cell.value for cell in worksheet[2]]   # 对第二行的每个单元格获取它的值，并将这些值放在一个列表中
        print("excel的列名为：",headers)
        # 结果： ['id', 'url', 'path', 'method', 'params', 'headers', 'data', 'type', 'actualResult', 'expectResult', 'result', 'caseName']

        # todo 遍历excel中的每一行数据，将excel的每一行数据当做value,存储为一个字典，并且放在data中
        data=[]   # data是需要执行的用例数据

        """
        1.worksheet.iter_rows() 用于按行迭代工作表中的数据
        2.min=3 指定迭代的起始行，因为第2行是列名，所以从第3行开始迭代
        3.values_only=True,指定只返回单元格中的值
        4.headers是列名列表，row是每一行的数据列表
        4.dict(zip(headers,row)),将headers列表中的元素作为键，row列表中的对应元素作为值，创建一个键值对应的字典
        """
        for row in worksheet.iter_rows(min_row=3,values_only=True):
            # 遍历每一行数据，将每一行数据当做字典存储
            new_data=dict(zip(headers,row))
            if new_data["is_true"] is True:
                # 用例过滤机制：如果excel文件中有用例不需要执行，可以添加一个字段为is_true,为True执行，为False时不执行,需要执行时加入到data列表中
                data.append(new_data)

        workbook.close()
        # data的数据格式是列表嵌套字典，每个字典对应一行数据，字典的key是列名，value是每一行对应的数据
        print("从excel读取的数据为：",data)
        return data



    @staticmethod
    #  TODO 写入excel文件，需要知道excel文件路径，sheet名、第几列，第几行、写入的值
    def write_excel(excel_file_path=EXCEL_FILE_PATH6,sheet_name=SHEET_NAME,column=None,row=None,value=None):
        # 打开已经存在的excel文件或者创建新的excel文件
        try:
            workbook=openpyxl.load_workbook(excel_file_path)
        except FileNotFoundError:
            workbook=openpyxl.Workbook()

        # 选择或者创建指定的工作表
        if sheet_name in workbook.sheetnames:
            worksheet=workbook[sheet_name]
        else:
            worksheet=workbook.create_sheet(sheet_name)

        # 写入数据到指定的行或者列.将获取到的value值写入到指定的行列中
        worksheet.cell(column=column,row=row).value=value

        # 写入完成后保存文件
        workbook.save(excel_file_path)

    @staticmethod
    def data_EncryptDataAES(data):
        """
        需要加密的接口，对于有@符号数据进行加密处理，没有@符号的数据不加密
        :param data: 文件中的data参数,{"@username":"tony","@password":"123456"}
        :return:
        """
        newdata = {}
        for key in data:
            if key[0]=="@":
                # 加密数据，ENCRYPTAES是在封装的加密方法Data_Encrypt_AES.py中已经创建好的实例化对象，可以直接调用方法进行加密
                # 加密后需要把key前面的@符号去掉，所以使用key[1:]去掉@符号
                newdata[key[1:]]=ENCRYPTAES.encrypt_data(data[key])
            else:
                # 如果key前面没有@符号，不加密直接赋值
                newdata[key]=data[key]

        return newdata










if __name__ == '__main__':
    # FileExcelRead.read_excel()

    # 需要加密处理：
    data={"@username":"tony","@password":"123456"}
    res=FileExcelRead.data_EncryptDataAES(data)
    print("加密后的数据：",res)

    # 不需要加密处理的数据：
    data={"username":"tony","password":"123456"}
    res=FileExcelRead.data_EncryptDataAES(data)
    print("不需要加密的数",res)

